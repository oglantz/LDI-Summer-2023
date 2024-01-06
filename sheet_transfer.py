import openpyxl
from openpyxl import Workbook
from pycel.excelcompiler import ExcelCompiler
from copy import copy
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from tqdm import tqdm
"""
EVERYTHING COMMENTED WE DON'T NEED! JUST THERE CAUSE IM SCARED OF DELETING CODE!
"""

class SheetReaderWriter:
    def __init__(self, result_file: str, file_list: list[str]):
        self._current_file = None
        self._current_book = None
        self._read_sheet = None
        self._eval_sheet = None
        self._file_list = file_list

        self._final_file = result_file
        self._final_book = Workbook()
        self._final_sheet = self._final_book.active
        self._file_count = 0

        self.ACCOUNTING_STYLE = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
        self.CURRENCY_STYLE = '"$"#,##0.00'
        self.overall_total = 0

    def cell_has_formatting(self, new_cell, old_cell, total = False):
        if total:
            new_cell.font = copy(old_cell.font)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)
            return
        else:
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)

    def return_cell(self, address):
        return self._current_book[self._read_sheet][address].value

    def _transfer_cells(self, start_cell, end_cell):
        for row in self._read_sheet[start_cell:end_cell]:
            for cell in row:
                new_cell = self._final_sheet.cell(row = cell.row + self._file_count,
                                                  column = cell.col_idx, value = cell.value)
                if cell.has_style:
                    self.cell_has_formatting(new_cell, cell)
                if cell.col_idx == 11:
                    if cell.number_format == self.CURRENCY_STYLE:
                        self.overall_total += cell.value
                    if cell.number_format == self.ACCOUNTING_STYLE:
                        self.overall_total += cell.value


    def _format_widths(self):
        ws = self._final_sheet
        dim_holder = DimensionHolder(worksheet = ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

        ws.column_dimensions = dim_holder

    def _display_total(self):
        total_label = self._final_sheet.cell(row = self._file_count + 1,
                               column = 2, value = "Overall Total:")
        sample_row = str(self._file_count - 4)
        sample_col = "B"
        self.cell_has_formatting(total_label, self._final_sheet[sample_col + sample_row], True)

        total_num = self._final_sheet.cell(row = self._file_count + 1,
                               column = 3, value = round(self.overall_total, 2))
        total_num.number_format = self.ACCOUNTING_STYLE

    def _save_result(self):
        self._final_book.save(self._final_file + ".xlsx")
        print(f"File saved as {self._final_file}.xlsx")

    def cycle_through_sheets(self):
        for file in tqdm(self._file_list):
            self._current_file = file
            self._current_book = openpyxl.load_workbook(file, data_only = True)
            self._read_sheet = self._current_book.active
            self._eval_sheet = self._current_book.sheetnames[0]
            self._transfer_cells('A1', 'L4')
            self._file_count += 7
        self._display_total()
        self._format_widths()
        self._save_result()
