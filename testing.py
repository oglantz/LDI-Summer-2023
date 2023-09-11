import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from copy import copy

read_from = load_workbook('IT LEAP Budget FY23-24 SPRING.xlsx')
read_sheet = read_from.active

write_to = Workbook()
write_sheet = write_to.active

for row in read_sheet['A1': 'L4']:
    for cell in row:
        new_cell = write_sheet.cell(row=cell.row, column=cell.col_idx,
                value= cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

write_to.save('test.xlsx')
